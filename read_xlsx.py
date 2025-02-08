from openpyxl import load_workbook

PATH = "format.xlsx" #Variable pour stoquer l'emplacement du fichier. il sera récupere dans un widget
NAME_PROMOTION = "Bac3 Info GL" #varible pour savoir la promotion gérer par le sécretaire du jury


from openpyxl import load_workbook

def excel_to_list_of_dictionaries(file_path, sheet_name=None):
    """
    Transforms an Excel sheet into a list of dictionaries to analyze the data.
    """
    wb = load_workbook(file_path)

    sheet_active = None
    if sheet_name:
        if sheet_name.lower() in [sheet.lower() for sheet in wb.sheetnames]:
            sheet_active = wb[sheet_name]
    else:
        sheet_active = wb.active  # Use active sheet if no name is provided

    if not sheet_active:
        raise ValueError("Sheet not found")

    header = []
    data = []
    
    for row_index, row in enumerate(sheet_active.iter_rows(), start=1):
        data_dict = {}
        for i, cell in enumerate(row):
            if row_index == 1:
                # First row: header
                header.append(cell.value.lower())
            else:
                data_dict[header[i]] = cell.value
        if row_index != 1:  # Skip adding an empty dictionary for the header row
            data.append(data_dict)

    return data

def excel_to_list_of_tuples(file_path, sheet_name=None):
    """
    Transforme une feuille Excel en une liste de tuples.
    """
    wb = load_workbook(file_path)

    sheet_active = None
    if sheet_name:
        if sheet_name.lower() in [sheet.lower() for sheet in wb.sheetnames]:
            sheet_active = wb[sheet_name]
    else:
        sheet_active = wb.active  # Utilise la feuille active si aucun nom n'est spécifié

    if not sheet_active:
        raise ValueError("Sheet not found")

    data = []

    for row_index, row in enumerate(sheet_active.iter_rows(values_only=True), start=1):
        if row_index == 1:
            continue  # Ignorer la première ligne (en-têtes)
        data.append(tuple(row))  # Convertir la ligne en tuple et l'ajouter à la liste

    return data

def filter_fonction(data):
	"""
	supprimer les données dont la matricule, le nom ou le prénom serait absente
	"""
	for i in data.copy():
		if i["matricule"]==None or i["nom"]==None or i["prenom"]==None:
			data.remove(i)
	return data

def filter_fonction_1(data):
    """
    Supprime les entrées où la matricule, le nom ou le prénom est absent (None ou vide).
    """
    return [i for i in data if i[0] and i[1] and i[2]]


def get_data_from_excel(file_path, sheet_name=None):
    
    data = excel_to_list_of_dictionaries(file_path, sheet_name)
    data = filter_fonction(data)
    
    return data

# print(get_data_from_excel("format.xlsx"))